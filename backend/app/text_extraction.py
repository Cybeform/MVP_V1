import fitz  # PyMuPDF
import docx
import pandas as pd
from openpyxl import load_workbook
from docx import Document as DocxDocument
import io
from typing import Optional

def extract_text_from_pdf(file_path: str) -> str:
    """Extraire le texte d'un fichier PDF"""
    try:
        doc = fitz.open(file_path)
        text = ""
        for page in doc:
            text += page.get_text()
        doc.close()
        return text
    except Exception as e:
        print(f"Erreur lors de l'extraction PDF: {e}")
        return ""

def extract_text_from_pdf_with_pages(file_path: str) -> str:
    """
    Extraire le texte d'un fichier PDF en incluant les numéros de page
    Chaque page est séparée par un marqueur spécial
    """
    try:
        doc = fitz.open(file_path)
        text = ""
        for page_num, page in enumerate(doc, 1):
            page_text = page.get_text()
            if page_text.strip():  # Seulement si la page contient du texte
                text += f"\n\n--- PAGE {page_num} ---\n\n"
                text += page_text
        doc.close()
        return text
    except Exception as e:
        print(f"Erreur lors de l'extraction PDF avec pages: {e}")
        return ""

def extract_text_from_docx(file_path: str) -> str:
    """Extraire le texte d'un fichier DOCX"""
    try:
        doc = docx.Document(file_path)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text
    except Exception as e:
        print(f"Erreur lors de l'extraction DOCX: {e}")
        return ""

def extract_text_from_xlsx(file_path: str) -> str:
    """Extraire le texte d'un fichier XLSX"""
    try:
        # Essayer avec pandas d'abord
        try:
            df = pd.read_excel(file_path, sheet_name=None)  # Lire toutes les feuilles
            text = ""
            for sheet_name, sheet_df in df.items():
                text += f"Feuille: {sheet_name}\n"
                text += sheet_df.to_string(index=False, na_rep='') + "\n\n"
            return text
        except:
            # Fallback avec openpyxl
            workbook = load_workbook(file_path, data_only=True)
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"Feuille: {sheet_name}\n"
                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
                text += "\n"
            return text
    except Exception as e:
        print(f"Erreur lors de l'extraction XLSX: {e}")
        return ""

def extract_text_from_file(file_path: str, file_type: str, include_pages: bool = False) -> str:
    """
    Extraire le texte d'un fichier selon son type
    
    Args:
        file_path: Chemin vers le fichier
        file_type: Type MIME du fichier
        include_pages: Si True, inclut les marqueurs de page pour les PDF
        
    Returns:
        Texte extrait du fichier
    """
    if file_type == "application/pdf":
        if include_pages:
            return extract_text_from_pdf_with_pages(file_path)
        else:
            return extract_text_from_pdf(file_path)
    elif file_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
        return extract_text_from_docx(file_path)
    elif file_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        return extract_text_from_xlsx(file_path)
    else:
        raise ValueError(f"Type de fichier non supporté: {file_type}")

def get_text_preview(text: str, max_length: int = 200) -> str:
    """Générer un aperçu du texte extrait"""
    if len(text) <= max_length:
        return text
    return text[:max_length] + "..." 